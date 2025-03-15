# conda activate allpy310

# pip install python-docx lxml openpyxl PyPDF2

# Преобразование XLSX в TXT
# Для работы с XLSX файлами используем библиотеку openpyxl.

from openpyxl import load_workbook

def xlsx_to_txt(xlsx_path, txt_path):
    wb = load_workbook(xlsx_path)
    with open(txt_path, 'w', encoding='utf-8') as txt_file:
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                txt_file.write('\t'.join(map(str, row)) + '\n')

# Пример использования
xlsx_to_txt('D:\\AI_DATA\\PDF_DOC_DOCX_TXT_XML\\ПК_189660-189760.xlsx', 'example.txt')
